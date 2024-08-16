import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime


def enviar_mensagem_whatsapp(nome, telefone, mensagem, imagem, tempo_espera):
    try:
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
        webbrowser.open(link_mensagem_whatsapp)
        sleep(tempo_espera)
        seta = pyautogui.locateCenterOnScreen(imagem)
        sleep(5)
        pyautogui.click(seta[0], seta[1])
        sleep(5)
        pyautogui.hotkey('ctrl', 'w')
        sleep(5)
        return True
    except pyautogui.ImageNotFoundException:
        print(f'Não foi possível encontrar a imagem para clicar em {nome}')
    except Exception as e:
        print(f'Não foi possível enviar mensagem para {nome}: {e}')
        with open('erros.csv', 'a', newline='', encoding='utf-8') as arquivo:
            arquivo.write(f'{nome},{telefone}{os.linesep}')
        return False


def processar_planilhas(filename, workbook_enviados, pagina_clientes, mensagem_padrao, imagem, tempo_espera,
                        continuar_execucao, horario_encerramento, paused):
    pagina_enviados = workbook_enviados.active
    pagina_enviados.title = 'Sheet1'

    # Copiar cabeçalho se a planilha _enviados estiver vazia
    if pagina_enviados.max_row == 1:
        for cell in pagina_clientes[1]:
            pagina_enviados[cell.column_letter + '1'] = cell.value

    row_to_delete = []

    for linha in pagina_clientes.iter_rows(min_row=2, values_only=False):
        if not continuar_execucao[0]:
            break
        if paused[0]:
            print("Programa pausado. Pressione 'c' para continuar ou 'q' para sair.")
            while paused[0]:
                sleep(1)
                if not continuar_execucao[0]:
                    break
            print("Continuando a execução...")
        if horario_encerramento and datetime.now() >= horario_encerramento:
            print("Horário de encerramento atingido. Encerrando o programa.")
            break
        nome = linha[0].value
        telefone = linha[1].value

        if nome is None or telefone is None:
            print("Encontrado valor None. Encerrando o programa.")
            break

        print(f'Enviando mensagem para {nome} ({telefone})...')
        sucesso = enviar_mensagem_whatsapp(nome, telefone, mensagem_padrao, imagem, tempo_espera)
        if sucesso:
            nova_linha = [cell.value for cell in linha]
            pagina_enviados.append(nova_linha)
            row_to_delete.append(linha[0].row)

    for row in sorted(row_to_delete, reverse=True):
        pagina_clientes.delete_rows(row)

    workbook_enviados.save(os.path.splitext(filename)[0] + '_enviados.xlsx')
    # Mova o salvamento do workbook principal para a função `main` após o término do processamento
    return True


def main():
    paused = [False]  # Use uma lista para que possa ser modificado dentro de funções internas
    continuar_execucao = [True]  # Use uma lista para que possa ser modificado dentro de funções internas
    workbook = None
    pagina_clientes = None
    mensagem_padrao = None
    imagem = None
    tempo_espera = 60  # Valor padrão de 60 segundos
    horario_encerramento = None  # Horário de encerramento do programa
    workbook_enviados = None
    filename = None

    def worker():
        nonlocal workbook
        nonlocal pagina_clientes
        nonlocal mensagem_padrao
        nonlocal imagem
        nonlocal tempo_espera
        nonlocal horario_encerramento
        nonlocal workbook_enviados
        nonlocal filename
        nonlocal continuar_execucao
        nonlocal paused

        if workbook is None or pagina_clientes is None or mensagem_padrao is None or imagem is None:
            messagebox.showerror("Erro",
                                 "Por favor, selecione o arquivo, insira a mensagem e selecione a imagem antes de iniciar.")
            return

        enviados_filename = os.path.splitext(filename)[0] + '_enviados.xlsx'
        if os.path.exists(enviados_filename):
            workbook_enviados = openpyxl.load_workbook(enviados_filename)
        else:
            workbook_enviados = openpyxl.Workbook()

        sucesso = processar_planilhas(filename, workbook_enviados, pagina_clientes, mensagem_padrao, imagem,
                                      tempo_espera, continuar_execucao, horario_encerramento, paused)

        if sucesso:
            workbook.save(filename)

        continuar_execucao[0] = False
        control_window.quit()

    thread = threading.Thread(target=worker)

    def pause_program():
        paused[0] = True

    def continue_program():
        paused[0] = False

    def quit_program():
        continuar_execucao[0] = False
        paused[0] = False
        control_window.quit()

    def select_file():
        nonlocal workbook
        nonlocal pagina_clientes
        nonlocal filename
        filename = filedialog.askopenfilename(initialdir="/", title="Selecione o arquivo",
                                              filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
        if filename:
            workbook = openpyxl.load_workbook(filename)
            pagina_clientes = workbook['Sheet1']
            print(f'Arquivo {filename} selecionado.')

    def input_message():
        nonlocal mensagem_padrao
        dialog = tk.Toplevel(root)
        dialog.title("Digite a mensagem")
        text = tk.Text(dialog, width=50, height=20)
        text.pack()
        ok_button = tk.Button(dialog, text="OK", command=lambda: set_message(text, dialog))
        ok_button.pack()

    def set_message(text_widget, dialog):
        nonlocal mensagem_padrao
        mensagem_padrao = text_widget.get("1.0", "end-1c")
        dialog.destroy()
        print('Mensagem definida.')

    def select_image():
        nonlocal imagem
        imagem = filedialog.askopenfilename(initialdir="/", title="Selecione a imagem",
                                            filetypes=(
                                            ("png files", "*.png"), ("jpeg files", "*.jpg"), ("all files", "*.*")))
        if imagem:
            print(f'Imagem {imagem} selecionada.')

    def start_program():
        nonlocal thread
        nonlocal tempo_espera
        nonlocal horario_encerramento
        try:
            tempo_espera = int(entry_tempo_espera.get())
        except ValueError:
            messagebox.showerror("Erro", "Por favor, insira um número válido para o tempo de espera.")
            return
        if entry_horario_encerramento.get():
            try:
                horario_encerramento = datetime.strptime(entry_horario_encerramento.get(), '%H:%M').replace(
                    year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
            except ValueError:
                messagebox.showerror("Erro", "Por favor, insira um horário válido no formato HH:MM.")
                return
        if workbook is None or pagina_clientes is None or mensagem_padrao is None or imagem is None:
            messagebox.showerror("Erro",
                                 "Por favor, selecione o arquivo, insira a mensagem e selecione a imagem antes de iniciar.")
            return
        root.destroy()
        control_window.deiconify()
        print('Iniciando o programa...')
        thread.start()

    root = tk.Tk()
    root.title("Selecionar Arquivo, Mensagem, Imagem e Tempo de Espera")

    select_file_button = tk.Button(root, text="Selecionar Arquivo", command=select_file)
    select_file_button.pack()

    input_message_button = tk.Button(root, text="Inserir Mensagem", command=input_message)
    input_message_button.pack()

    select_image_button = tk.Button(root, text="Selecionar Imagem", command=select_image)
    select_image_button.pack()

    tk.Label(root, text="Tempo de Espera (segundos):").pack()
    entry_tempo_espera = tk.Entry(root)
    entry_tempo_espera.insert(0, "60")
    entry_tempo_espera.pack()

    tk.Label(root, text="Horário de Encerramento (HH:MM):").pack()
    entry_horario_encerramento = tk.Entry(root)
    entry_horario_encerramento.pack()

    start_button = tk.Button(root, text="Iniciar", command=start_program)
    start_button.pack()

    control_window = tk.Tk()
    control_window.title("Controle do Programa")
    control_window.withdraw()

    pause_button = tk.Button(control_window, text="Pausar", command=pause_program)
    pause_button.pack()

    continue_button = tk.Button(control_window, text="Continuar", command=continue_program)
    continue_button.pack()

    quit_button = tk.Button(control_window, text="Sair", command=quit_program)
    quit_button.pack()

    root.mainloop()

    while continuar_execucao[0]:
        control_window.update()
        sleep(0.1)

    print('Programa encerrado.')


if __name__ == "__main__":
    main()
